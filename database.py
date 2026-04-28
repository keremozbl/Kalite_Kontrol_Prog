# -*- coding: utf-8 -*-
"""
Endüstriyel Kalite Kontrol Sistemi — Veritabanı Modülü
========================================================
SQLite3 ile üretim kayıtlarının yönetimi, sorgulama ve
Excel dışa aktarım işlemleri.

Tablo: uretim_kayitlari
Kolonlar: id, seri_no, durum, beyaz_kece, gri_kece,
          bakir_halka_tespit, hata_detayi, gorsel_yolu, tarih_saat
"""

import os
import sqlite3
import logging
from datetime import datetime
from typing import List, Tuple, Optional, Dict, Any

logger = logging.getLogger("database")


class DatabaseManager:
    """SQLite veritabanı yöneticisi — thread-safe singleton."""

    _instance = None

    def __new__(cls, db_path: str = None):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self, db_path: str = None):
        if self._initialized:
            return
        from config import DB_PATH
        self.db_path = db_path or DB_PATH
        self._init_db()
        self._initialized = True
        logger.info(f"Veritabanı başlatıldı: {self.db_path}")

    def _get_connection(self) -> sqlite3.Connection:
        """Thread-safe bağlantı oluştur."""
        conn = sqlite3.connect(self.db_path, timeout=10)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")  # Eşzamanlı okuma/yazma
        conn.execute("PRAGMA busy_timeout=5000")
        return conn

    def _init_db(self):
        """Veritabanı tablolarını oluştur."""
        with self._get_connection() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS uretim_kayitlari (
                    id              INTEGER PRIMARY KEY AUTOINCREMENT,
                    seri_no         TEXT    DEFAULT '',
                    durum           TEXT    NOT NULL CHECK(durum IN ('OK', 'NOK')),
                    beyaz_kece      INTEGER NOT NULL DEFAULT 0,
                    gri_kece        INTEGER NOT NULL DEFAULT 0,
                    bakir_halka_tespit INTEGER NOT NULL DEFAULT 0,
                    hata_detayi     TEXT    DEFAULT '',
                    gorsel_yolu     TEXT    DEFAULT '',
                    tarih_saat      TEXT    NOT NULL DEFAULT (datetime('now', 'localtime'))
                )
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_tarih
                ON uretim_kayitlari(tarih_saat)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_durum
                ON uretim_kayitlari(durum)
            """)
            conn.commit()
        logger.info("Veritabanı tabloları hazır.")

    # ─────────────────── KAYIT İŞLEMLERİ ─────────────────────────────

    def kayit_ekle(
        self,
        seri_no: str,
        durum: str,
        beyaz_kece: bool,
        gri_kece: bool,
        bakir_halka_tespit: bool = False,
        hata_detayi: str = "",
        gorsel_yolu: str = ""
    ) -> int:
        """
        Yeni üretim kaydı ekle.

        Args:
            seri_no: Okunan seri numarası
            durum: "OK" veya "NOK"
            beyaz_kece: Beyaz keçe var mı (True/False)
            gri_kece: Gri keçe var mı (True/False)
            bakir_halka_tespit: Bakır halka görüldü mü (True = keçe eksik)
            hata_detayi: Hata açıklaması
            gorsel_yolu: NOK görüntü dosya yolu

        Returns:
            Eklenen kaydın ID'si
        """
        tarih_saat = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with self._get_connection() as conn:
            cursor = conn.execute("""
                INSERT INTO uretim_kayitlari
                    (seri_no, durum, beyaz_kece, gri_kece,
                     bakir_halka_tespit, hata_detayi, gorsel_yolu, tarih_saat)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                seri_no, durum,
                int(beyaz_kece), int(gri_kece),
                int(bakir_halka_tespit),
                hata_detayi, gorsel_yolu, tarih_saat
            ))
            conn.commit()
            kayit_id = cursor.lastrowid
        logger.info(
            f"Kayıt #{kayit_id}: {durum} | Seri: {seri_no} | "
            f"Beyaz: {beyaz_kece} | Gri: {gri_kece} | Bakır: {bakir_halka_tespit}"
        )
        return kayit_id

    # ─────────────────── SORGU İŞLEMLERİ ─────────────────────────────

    def son_kayitlar(self, n: int = 10) -> List[Dict[str, Any]]:
        """Son N üretim kaydını getir (en yeni ilk)."""
        with self._get_connection() as conn:
            rows = conn.execute(
                "SELECT * FROM uretim_kayitlari ORDER BY id DESC LIMIT ?",
                (n,)
            ).fetchall()
        return [dict(row) for row in rows]

    def tarih_araliginda_getir(
        self,
        baslangic: str,
        bitis: str
    ) -> List[Dict[str, Any]]:
        """
        Tarih aralığında kayıtları getir.

        Args:
            baslangic: "YYYY-MM-DD" veya "YYYY-MM-DD HH:MM:SS"
            bitis: "YYYY-MM-DD" veya "YYYY-MM-DD HH:MM:SS"
        """
        with self._get_connection() as conn:
            rows = conn.execute(
                """SELECT * FROM uretim_kayitlari
                   WHERE tarih_saat BETWEEN ? AND ?
                   ORDER BY id DESC""",
                (baslangic, bitis)
            ).fetchall()
        return [dict(row) for row in rows]

    def istatistikler(self) -> Dict[str, int]:
        """Toplam, OK ve NOK sayaçlarını döndür."""
        with self._get_connection() as conn:
            toplam = conn.execute(
                "SELECT COUNT(*) FROM uretim_kayitlari"
            ).fetchone()[0]
            ok_sayi = conn.execute(
                "SELECT COUNT(*) FROM uretim_kayitlari WHERE durum='OK'"
            ).fetchone()[0]
            nok_sayi = conn.execute(
                "SELECT COUNT(*) FROM uretim_kayitlari WHERE durum='NOK'"
            ).fetchone()[0]
        return {
            "toplam": toplam,
            "ok": ok_sayi,
            "nok": nok_sayi,
            "oran": round(ok_sayi / toplam * 100, 1) if toplam > 0 else 0.0
        }

    def gunluk_istatistik(self) -> Dict[str, int]:
        """Bugünün istatistiklerini döndür."""
        bugun = datetime.now().strftime("%Y-%m-%d")
        with self._get_connection() as conn:
            toplam = conn.execute(
                "SELECT COUNT(*) FROM uretim_kayitlari WHERE tarih_saat LIKE ?",
                (f"{bugun}%",)
            ).fetchone()[0]
            ok_sayi = conn.execute(
                "SELECT COUNT(*) FROM uretim_kayitlari WHERE durum='OK' AND tarih_saat LIKE ?",
                (f"{bugun}%",)
            ).fetchone()[0]
            nok_sayi = conn.execute(
                "SELECT COUNT(*) FROM uretim_kayitlari WHERE durum='NOK' AND tarih_saat LIKE ?",
                (f"{bugun}%",)
            ).fetchone()[0]
        return {
            "toplam": toplam,
            "ok": ok_sayi,
            "nok": nok_sayi,
            "oran": round(ok_sayi / toplam * 100, 1) if toplam > 0 else 0.0
        }

    # ─────────────────── EXCEL DIŞA AKTARIM ──────────────────────────

    def excel_aktar(
        self,
        dosya_yolu: str,
        baslangic: str = None,
        bitis: str = None
    ) -> str:
        """
        Üretim kayıtlarını Excel (.xlsx) dosyasına aktar.

        Args:
            dosya_yolu: Çıktı dosya yolu
            baslangic: Başlangıç tarihi (opsiyonel)
            bitis: Bitiş tarihi (opsiyonel)

        Returns:
            Oluşturulan dosya yolu
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error("openpyxl yüklü değil: pip install openpyxl")
            raise ImportError("Excel dışa aktarım için openpyxl gereklidir.")

        # Verileri getir
        if baslangic and bitis:
            kayitlar = self.tarih_araliginda_getir(baslangic, bitis)
        else:
            with self._get_connection() as conn:
                rows = conn.execute(
                    "SELECT * FROM uretim_kayitlari ORDER BY id DESC"
                ).fetchall()
            kayitlar = [dict(row) for row in rows]

        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Üretim Kayıtları"

        # Başlık stili
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Başlıklar
        basliklar = [
            "ID", "Seri Numarası", "Durum",
            "Beyaz Keçe", "Gri Keçe", "Bakır Halka (Eksik Keçe)",
            "Hata Detayı", "Tarih-Saat"
        ]
        for col, baslik in enumerate(basliklar, 1):
            cell = ws.cell(row=1, column=col, value=baslik)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Veri satırları
        ok_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        nok_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

        for row_idx, kayit in enumerate(kayitlar, 2):
            values = [
                kayit["id"],
                kayit["seri_no"],
                kayit["durum"],
                "✓ Var" if kayit["beyaz_kece"] else "✗ Yok",
                "✓ Var" if kayit["gri_kece"] else "✗ Yok",
                "⚠ Tespit" if kayit["bakir_halka_tespit"] else "-",
                kayit["hata_detayi"],
                kayit["tarih_saat"],
            ]
            fill = ok_fill if kayit["durum"] == "OK" else nok_fill
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sütun genişlikleri
        genislikler = [6, 20, 10, 14, 14, 22, 30, 22]
        for i, w in enumerate(genislikler, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # Kaydet
        wb.save(dosya_yolu)
        logger.info(f"Excel dışa aktarıldı: {dosya_yolu} ({len(kayitlar)} kayıt)")
        return dosya_yolu

    # ─────────────────── TEMİZLİK ────────────────────────────────────

    def sayaclari_sifirla(self):
        """Tüm kayıtları sil (dikkatli kullanın)."""
        with self._get_connection() as conn:
            conn.execute("DELETE FROM uretim_kayitlari")
            conn.execute("DELETE FROM sqlite_sequence WHERE name='uretim_kayitlari'")
            conn.commit()
        logger.warning("Tüm üretim kayıtları silindi!")

    def close(self):
        """Singleton'ı sıfırla."""
        DatabaseManager._instance = None
        self._initialized = False
        logger.info("Veritabanı bağlantısı kapatıldı.")
